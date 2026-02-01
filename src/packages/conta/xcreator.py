""" Excel simple creator.

Uses openpyxl !
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .wcell import GenCell


class ExcelGera(GenCell):
    """ Excel generator. """
    numeric_fmt = "0.00"

    def __init__(self, data, outname="out.xlsx", name=""):
        super().__init__(data=data if data else [], name=name if name else outname)
        self.outname = outname

    def generate(self):
        wbk = Workbook()
        ws = wbk.active
        ws.title = "Folha1"
        for row in self._data:
            ws.append(row)
        self._adjust_widths(ws)
        wbk.save(self.outname)

    def _best_cell_form(self, cell):
        """ Format cell basically """
        if cell.value is None:
            return None
        cell.number_format = "0.00"
        return cell

    def _adjust_widths(self, ws):
        """ Larguras de colunas. """
        pigment = 4
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
                if cell.value is not None and isinstance(cell.value, (float, int)):
                    if isinstance(ExcelGera.numeric_fmt, str):
                        cell.number_format = ExcelGera.numeric_fmt
            ws.column_dimensions[col_letter].width = max_len + pigment

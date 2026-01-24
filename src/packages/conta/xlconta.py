""" Excel simplifier.

Extracts Excel into simpler manageable fields.

Uses openpyxl !
"""

from openpyxl.utils import column_index_from_string, get_column_letter
from .wcell import WCell
from .xwordwrap import ascii_7bit

DEBUG = 0


class GenericConta:
    """ Generic Accounting class. """
    def __init__(self, data=None, flt=None, name="Conta"):
        self.name = name
        self._filter = flt
        data = [] if data is None else data
        assert isinstance(data, list), f"Not a list: {name}"
        self._data = data

    def get_all(self):
        """ Returns the data """
        assert self._data, self.name
        return self._data


class ExBook(GenericConta):
    """ Linear Workbook """
    def __init__(self, workbook, sheet_index=0, flt=None, name="Excel"):
        """
        *	add_empties is True if you want to add empty lines
        in case row gets filtered out; redefine class if that is the case.
        """
        super().__init__(flt=flt, name=name)
        self._filter = self._filter_prep(flt)
        self._add_empties = False
        self._data = self._linearize(workbook, sheet_index)

    def get_rows(self, sheet_index=1, row_start=0):
        """ sheet_index >= 1 """
        res = self._data[sheet_index]
        for n_row, row in res:
            use = row_start == 0 or (row_start >= 1 and n_row >= row_start)
            if not use:
                continue
            a_row = row
            yield a_row

    def get_dicts(self, sheet_index=1, row_start=0):
        """ Same as get_rows(), but returns a list of dictionaries
        containing column values.
        """
        res = self._data[sheet_index]
        for n_row, row in res:
            use = row_start == 0 or (row_start >= 1 and n_row >= row_start)
            if not use:
                continue
            d_row = {}
            for idx, val in enumerate(row, 1):
                col = get_column_letter(idx)
                d_row[col] = val
            yield d_row

    def _linearize(self, workbook, sheet_index):
        """ Simple linearization of Excel rows. """
        this = [self.name]	# Leading sheet as index 0
        data = [this]
        for idx, sheet in enumerate(workbook.worksheets, 1):
            this = []
            use = sheet_index <= 0 or idx == sheet_index
            if not use:
                continue
            for r_num, row in enumerate(
                sheet.iter_rows(values_only=True),
                1,
            ):
                a_row = self._adder(row, r_num, self._filter)
                use = a_row or self._add_empties
                if not use:
                    continue
                this.append(
                    (
                        r_num,
                        to_list(a_row),
                    )
                )
            data.append(this)
        return data

    def _adder(self, row, r_num, flt):
        """ res = list(row) --> max linear!
        """
        res = [
            (get_column_letter(idx), val)
            for idx, val in enumerate(row, 1)
        ]
        if flt is None:
            return res
        now = self._get_row_from_filters(
            res,
            r_num,
            flt,
        )
        return now

    def _get_row_from_filters(self, lst, r_num, flt):
        res = []
        for trip in flt:
            use, this = self._from_filter(lst, r_num, trip)
            #print(":::", r_num, use, trip, lst[:3])
            if use:
                return this
        return res

    def _from_filter(self, lst, r_num, trip):
        """ Processes one filter! """
        col_idx, oper, val = trip
        try:
            cell = lst[col_idx]
        except IndexError as err:
            print(f"{self.name}:row {r_num}: {err}")
            return None
        if cell is None:
            return False, []
        name = f"{get_column_letter(col_idx)}{r_num}"
        new = WCell(cell, name=name)
        #print(f"::: CHECK ({new.name}):", r_num, str(cell), [new], trip)
        if oper in ("=*",):
            if val in str(cell).lower():
                return True, lst
        elif oper in ("=",):
            if new.lower().startswith(val):
                return True, lst
        return False, []

    def _filter_prep(self, flt):
        """ Preparar o filtro """
        if not flt:
            return None
        res = []
        for trip in flt:
            col_letter, oper, val = trip
            col_idx = column_index_from_string(col_letter) - 1
            val = val.lower()
            res.append((col_idx, oper, val))
        return res


def easier(cell, col="@", debug=DEBUG):
    """ Gets a cell and a column letter. """
    if debug > 0:
        astr = f"{cell} (type={type(cell)})"
    else:
        astr = ascii_7bit(cell)
    return astr


def to_list(row):
    res = [
        (col, WCell(cell).to_string()) for col, cell in row
    ]
    return res
